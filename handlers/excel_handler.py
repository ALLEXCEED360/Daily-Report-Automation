# handlers/excel_handler.py
import os
import openpyxl
from datetime import datetime
from typing import Optional, Union

class ExcelHandler:
    """
    Handles loading the Excel template, picking the sheet for a day number,
    writing fields, and saving safely (atomic replace with fallback).
    """
    def __init__(self, template_path: Optional[str] = None):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        # Default to project-root/Others/daily_report_template.xlsx
        if template_path is None:
            template_path = os.path.join(base_dir, "..", "Others", "daily_report_template.xlsx")
        self.template_path = os.path.abspath(template_path)
        self.workbook = None
        self.sheet = None
        self.current_day = None

    def load_template(self, specific_day: Union[int, str, None] = None) -> bool:
        """Load workbook and select a sheet named by day number (e.g., '17')."""
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template not found at: {self.template_path}")

            self.workbook = openpyxl.load_workbook(self.template_path)

            # Determine sheet name
            if specific_day is None:
                self.current_day = str(datetime.now().day)
            else:
                self.current_day = str(specific_day)

            # Find sheet
            sheet_found = False
            for s in self.workbook.sheetnames:
                if str(s) == self.current_day:
                    self.sheet = self.workbook[s]
                    sheet_found = True
                    break

            if sheet_found:
                print(f"✓ Template loaded successfully - Using sheet '{self.current_day}'")
                return True
            else:
                print(f"✗ Could not find sheet '{self.current_day}' in workbook")
                print(f"Available sheets: {self.workbook.sheetnames}")
                return False

        except Exception as e:
            print(f"✗ Error loading template: {e}")
            return False

    def list_all_sheets(self):
        if self.workbook:
            print("Available sheets:")
            for i, s in enumerate(self.workbook.sheetnames, 1):
                print(f"  {i}. {s}")
        else:
            print("No workbook loaded")

    def update_date(self, date_obj: Optional[Union[str, datetime]] = None) -> str:
        """Write a formatted date to cell E3. Accepts string 'YYYY-MM-DD' or datetime."""
        if self.sheet is None:
            raise RuntimeError("No sheet loaded. Call load_template() first.")
        if date_obj is None:
            use_dt = datetime.now()
        elif isinstance(date_obj, str):
            # Try common formats
            parsed = None
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    parsed = datetime.strptime(date_obj, fmt)
                    break
                except Exception:
                    continue
            if parsed is None:
                raise ValueError("Unrecognized date string format.")
            use_dt = parsed
        elif isinstance(date_obj, datetime):
            use_dt = date_obj
        else:
            raise ValueError("Unsupported date_obj type")

        date_str = use_dt.strftime("%m/%d/%Y")
        self.sheet['E3'] = date_str
        print(f"✓ Date updated to: {date_str}")
        return date_str

    def save_report(self, filename: Optional[str] = None, make_backup: bool = True) -> bool:
        """
        Save workbook.
        Default: overwrite self.template_path; create backup first if make_backup True.
        Strategy: write to tmp file and `os.replace()` to attempt atomic overwrite.
        If target is locked, save a timestamped copy instead.
        """
        if self.workbook is None:
            print("✗ No workbook to save.")
            return False

        target = os.path.abspath(filename) if filename else self.template_path
        target_dir = os.path.dirname(target)
        tmp_target = os.path.join(target_dir, f".{os.path.basename(target)}.tmp")
        timestamped = target.replace(".xlsx", f".saved_{datetime.now().strftime('%Y%m%dT%H%M%S')}.xlsx")

        try:
            if filename is None and make_backup:
                try:
                    backup_name = target.replace(".xlsx", f".backup_{datetime.now().strftime('%Y%m%dT%H%M%S')}.xlsx")
                    self.workbook.save(backup_name)
                    print(f"✓ Backup saved as: {backup_name}")
                except Exception as be:
                    print(f"⚠️ Could not create backup: {be} (continuing)")

            # Try tmp then replace
            try:
                self.workbook.save(tmp_target)
                try:
                    os.replace(tmp_target, target)
                    print(f"✓ Report saved as: {target} (atomic replace)")
                    return True
                except PermissionError as perr:
                    # locked: remove tmp and fallback
                    try:
                        os.remove(tmp_target)
                    except Exception:
                        pass
                    raise perr

            except PermissionError as perm_err:
                print(f"⚠️ PermissionError when attempting overwrite: {perm_err}")
                try:
                    self.workbook.save(timestamped)
                    print(f"✓ Could not overwrite; saved copy as: {timestamped}")
                    return True
                except Exception as e2:
                    print(f"✗ Failed to save fallback copy: {e2}")
                    return False

            except Exception as e:
                print(f"✗ Error saving temporary file: {e}")
                try:
                    self.workbook.save(timestamped)
                    print(f"✓ Saved copy as: {timestamped}")
                    return True
                except Exception as e3:
                    print(f"✗ Also failed saving fallback file: {e3}")
                    return False

        except Exception as final_e:
            print(f"✗ Unexpected error in save_report: {final_e}")
            return False

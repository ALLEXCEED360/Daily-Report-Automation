import openpyxl
from datetime import datetime

class ExcelHandler:
    def __init__(self, template_path="Others\daily_report_template.xlsx"):
        """Initialize the Excel handler with your template"""
        self.template_path = template_path
        self.workbook = None
        self.sheet = None
        self.current_day = None
        
    def load_template(self, specific_day=None):
        """Load the Excel template and select the correct sheet"""
        try:
            self.workbook = openpyxl.load_workbook(self.template_path)
            
            # Get today's day or use specific day
            if specific_day:
                self.current_day = str(specific_day)
            else:
                self.current_day = str(datetime.now().day)
            
            # Try to find the sheet with the day number
            sheet_found = False
            for sheet_name in self.workbook.sheetnames:
                if str(sheet_name) == self.current_day:
                    self.sheet = self.workbook[sheet_name]
                    sheet_found = True
                    break
            
            if sheet_found:
                print(f"✓ Template loaded successfully - Using sheet '{self.current_day}'")
                return True
            else:
                print(f"✗ Could not find sheet '{self.current_day}' in the workbook")
                print(f"Available sheets: {self.workbook.sheetnames}")
                return False
                
        except Exception as e:
            print(f"✗ Error loading template: {e}")
            return False
    
    def list_all_sheets(self):
        """List all available sheets in the workbook"""
        if self.workbook:
            print("Available sheets:")
            for i, sheet_name in enumerate(self.workbook.sheetnames, 1):
                print(f"  {i}. {sheet_name}")
        else:
            print("No workbook loaded")
    
    def update_date(self, date_str=None):
        """Update the date field"""
        if date_str is None:
            date_str = datetime.now().strftime("%m/%d/%Y")
        
        self.sheet['E3'] = date_str
        print(f"✓ Date updated to: {date_str}")
    
    def update_lotto_data(self, net_sales=None, start_num=None, end_num=None):
        """Update lotto-related fields"""
        if net_sales:
            self.sheet['T30'] = net_sales  # Lotto M/C Net Sales
            print(f"✓ Lotto Net Sales: {net_sales}")
        
        if start_num:
            print(f"✓ Lotto Start: {start_num}")
        
        if end_num:
            print(f"✓ Lotto End: {end_num}")
    
    def save_report(self, filename=None):
        """Save the updated report"""
        if filename is None:
            today = datetime.now().strftime("%Y%m%d")
            filename = f"daily_report_{today}.xlsx"
        
        try:
            self.workbook.save(filename)
            print(f"✓ Report saved as: {filename}")
            return True
        except Exception as e:
            print(f"✗ Error saving report: {e}")
            return False

# Test the Excel handler
if __name__ == "__main__":
    print("Testing Excel Handler with Multiple Sheets...")
    
    # Create handler and test it
    excel = ExcelHandler()
    
    # First, let's see all available sheets
    if excel.workbook is None:
        excel.workbook = openpyxl.load_workbook("daily_report_template.xlsx")
    
    excel.list_all_sheets()
    
    # Now try to load the correct sheet
    if excel.load_template():
        excel.update_date()
        excel.update_lotto_data(net_sales=1250.75)
        excel.save_report("test_report_day17.xlsx")
        print(f"\n✓ Excel handler test completed successfully for day {excel.current_day}!")
    else:
        print("✗ Could not load the correct sheet")